import csv
import io
import re
from pathlib import Path
from typing import Any, Optional, TypeVar
from urllib.parse import unquote, urlparse
from uuid import uuid4

import httpx
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.config import get_settings
from app.database import get_db
from app.json_utils import dumps, loads
from app.models import AuditRule, AuditTask, DetectionItem, ImportTask, Industry, Lab, ModelProvider, Quote, Standard, StandardClause
from app.services.clause_chunker import chunk_standard_text
from app.services.document_parser import get_document_parser
from app.services.model_gateway import get_model_gateway
from app.services.tools import get_tool_statuses, test_tool
from app.schemas import (
    AuditRuleCreate,
    AuditRuleRead,
    DashboardSummary,
    DetectionItemCreate,
    DetectionItemRead,
    IndustryCreate,
    IndustryRead,
    ImportTaskRead,
    LabCreate,
    LabRead,
    ModelProviderCreate,
    ModelProviderRead,
    StandardCreate,
    StandardClauseRead,
    StandardRead,
)

router = APIRouter(prefix="/api/admin", tags=["admin"])
ModelT = TypeVar("ModelT")


def _get_or_404(db: Session, model: type[ModelT], item_id: str) -> ModelT:
    item = db.get(model, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    return item


def _is_pending_effective_standard(standard: Standard) -> bool:
    clauses = loads(standard.clauses, [])
    text = f"{standard.name} {standard.version} {standard.effective_date} {dumps(clauses)}"
    return any(keyword in text for keyword in ["待实施", "过渡", "已发布待实施", "2027-03-16", "2026-03-16"])


def _standard_quality_score(item: Standard) -> int:
    clauses = loads(item.clauses, [])
    return (
        len(clauses) * 5
        + (20 if item.source_file and Path(item.source_file).exists() else 0)
        + (10 if item.effective_date else 0)
        + (6 if item.version and item.version != "现行" else 0)
        + (4 if item.status == "active" else 0)
        + len(item.name)
    )


def _find_standard_by_code(db: Session, industry_id: str, code: str) -> Optional[Standard]:
    normalized = code.strip().upper().replace(" ", "")
    for item in db.scalars(select(Standard).where(Standard.industry_id == industry_id)):
        if item.code.strip().upper().replace(" ", "") == normalized:
            return item
    return None


def _ensure_clause_table(db: Session) -> None:
    StandardClause.__table__.create(bind=db.get_bind(), checkfirst=True)


def _replace_standard_clauses(
    db: Session,
    standard: Standard,
    text: str,
    source_file: str,
) -> int:
    _ensure_clause_table(db)
    for existing in db.scalars(select(StandardClause).where(StandardClause.standard_id == standard.id)):
        db.delete(existing)
    chunks = chunk_standard_text(text)
    for index, chunk in enumerate(chunks, 1):
        db.add(
            StandardClause(
                standard_id=standard.id,
                industry_id=standard.industry_id,
                clause_no=chunk.clause_no[:80],
                title=chunk.title[:240],
                content=chunk.content,
                page_no=chunk.page_no[:40],
                source_file=source_file,
                chunk_index=index,
                status="active",
            )
        )
    standard.clauses = dumps(
        [
            {"no": chunk.clause_no, "title": chunk.title, "content": chunk.content[:500]}
            for chunk in chunks[:20]
        ]
    )
    return len(chunks)


@router.get("/dashboard", response_model=DashboardSummary)
def dashboard(db: Session = Depends(get_db)) -> DashboardSummary:
    return DashboardSummary(
        industry_count=db.scalar(select(func.count()).select_from(Industry)) or 0,
        standard_count=db.scalar(select(func.count()).select_from(Standard)) or 0,
        rule_count=db.scalar(select(func.count()).select_from(AuditRule)) or 0,
        detection_item_count=db.scalar(select(func.count()).select_from(DetectionItem)) or 0,
        audit_task_count=db.scalar(select(func.count()).select_from(AuditTask)) or 0,
        quote_count=db.scalar(select(func.count()).select_from(Quote)) or 0,
        needs_review_count=db.scalar(select(func.count()).select_from(AuditTask).where(AuditTask.needs_human_review.is_(True))) or 0,
    )


@router.get("/tools/status")
def tools_status(db: Session = Depends(get_db)) -> dict[str, Any]:
    return get_tool_statuses(db)


@router.post("/tools/{tool_id}/test")
def test_admin_tool(tool_id: str, db: Session = Depends(get_db)) -> dict[str, Any]:
    return test_tool(db, tool_id)


@router.get("/industries", response_model=list[IndustryRead])
def list_industries(db: Session = Depends(get_db)) -> list[Industry]:
    return list(db.scalars(select(Industry).order_by(Industry.created_at)))


@router.post("/industries", response_model=IndustryRead)
def create_industry(payload: IndustryCreate, db: Session = Depends(get_db)) -> Industry:
    item = Industry(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.patch("/industries/{item_id}", response_model=IndustryRead)
def update_industry(item_id: str, payload: IndustryCreate, db: Session = Depends(get_db)) -> Industry:
    item = _get_or_404(db, Industry, item_id)
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return item


@router.delete("/industries/{item_id}", response_model=IndustryRead)
def delete_industry(item_id: str, db: Session = Depends(get_db)) -> Industry:
    item = _get_or_404(db, Industry, item_id)
    item.status = "inactive"
    db.commit()
    db.refresh(item)
    return item


@router.get("/standards", response_model=list[StandardRead])
def list_standards(db: Session = Depends(get_db)) -> list[StandardRead]:
    items = list(db.scalars(select(Standard).order_by(Standard.created_at.desc())))
    return [StandardRead.model_validate(item, from_attributes=True).model_copy(update={"clauses": loads(item.clauses, [])}) for item in items]


@router.get("/knowledge/coverage")
def knowledge_coverage(db: Session = Depends(get_db)) -> dict[str, Any]:
    _ensure_clause_table(db)
    industries = list(db.scalars(select(Industry).order_by(Industry.created_at)))
    standards = list(db.scalars(select(Standard)))
    rules = list(db.scalars(select(AuditRule).where(AuditRule.status == "active")))
    clause_counts = dict(
        db.execute(
            select(StandardClause.industry_id, func.count(StandardClause.id))
            .where(StandardClause.status == "active")
            .group_by(StandardClause.industry_id)
        ).all()
    )
    by_industry: list[dict[str, Any]] = []
    for industry in industries:
        industry_standards = [item for item in standards if item.industry_id == industry.id and item.status == "active"]
        industry_rules = [item for item in rules if item.industry_id == industry.id]
        pending = [item for item in industry_standards if _is_pending_effective_standard(item)]
        source_ready = [item for item in industry_standards if item.source_file and Path(item.source_file).exists()]
        by_industry.append(
            {
                "industry_id": industry.id,
                "code": industry.code,
                "name": industry.name,
                "standard_count": len(industry_standards),
                "rule_count": len(industry_rules),
                "clause_count": int(clause_counts.get(industry.id, 0)),
                "pending_effective_count": len(pending),
                "source_ready_count": len(source_ready),
                "sample_standards": [
                    {
                        "code": item.code,
                        "name": item.name,
                        "version": item.version,
                        "effective_date": item.effective_date,
                    }
                    for item in sorted(industry_standards, key=lambda value: value.code)[:6]
                ],
            }
        )
    return {
        "total_standards": sum(item["standard_count"] for item in by_industry),
        "total_rules": sum(item["rule_count"] for item in by_industry),
        "total_clauses": sum(item["clause_count"] for item in by_industry),
        "source_ready_count": sum(item["source_ready_count"] for item in by_industry),
        "industries": by_industry,
        "note": "内置法规包为目录、摘要和审核要点；官方全文可通过知识库导入补充。",
    }


@router.get("/standards/{item_id}/clauses", response_model=list[StandardClauseRead])
def list_standard_clauses(item_id: str, db: Session = Depends(get_db)) -> list[StandardClause]:
    _ensure_clause_table(db)
    _get_or_404(db, Standard, item_id)
    return list(
        db.scalars(
            select(StandardClause)
            .where(StandardClause.standard_id == item_id, StandardClause.status == "active")
            .order_by(StandardClause.chunk_index)
        )
    )


@router.get("/knowledge/clauses/search", response_model=list[StandardClauseRead])
def search_standard_clauses(
    q: str = Query(""),
    industry_id: Optional[str] = Query(None),
    limit: int = Query(20, ge=1, le=80),
    db: Session = Depends(get_db),
) -> list[StandardClause]:
    _ensure_clause_table(db)
    statement = select(StandardClause).where(StandardClause.status == "active")
    if industry_id:
        statement = statement.where(StandardClause.industry_id == industry_id)
    if q.strip():
        keyword = f"%{q.strip()}%"
        statement = statement.where(
            StandardClause.content.like(keyword)
            | StandardClause.title.like(keyword)
            | StandardClause.clause_no.like(keyword)
        )
    return list(db.scalars(statement.order_by(StandardClause.created_at.desc(), StandardClause.chunk_index).limit(limit)))


@router.post("/knowledge/dedupe-standards")
def dedupe_standards(db: Session = Depends(get_db), hard: bool = Query(False)) -> dict[str, Any]:
    standards = list(db.scalars(select(Standard).order_by(Standard.created_at.desc())))
    groups: dict[tuple[str, str], list[Standard]] = {}
    for item in standards:
        normalized_code = item.code.strip().upper().replace(" ", "")
        groups.setdefault((item.industry_id, normalized_code), []).append(item)

    duplicate_groups = [items for items in groups.values() if len(items) > 1]
    removed = 0
    relinked_rules = 0
    inactive = 0
    kept: list[dict[str, str]] = []
    for items in duplicate_groups:
        winner = max(items, key=_standard_quality_score)
        kept.append({"id": winner.id, "code": winner.code, "name": winner.name})
        for duplicate in items:
            if duplicate.id == winner.id:
                continue
            for rule in db.scalars(select(AuditRule).where(AuditRule.standard_id == duplicate.id)):
                rule.standard_id = winner.id
                relinked_rules += 1
            if hard:
                db.delete(duplicate)
                removed += 1
            else:
                duplicate.status = "inactive"
                inactive += 1
    db.commit()
    return {
        "duplicate_groups": len(duplicate_groups),
        "inactive": inactive,
        "removed": removed,
        "relinked_rules": relinked_rules,
        "kept": kept[:20],
        "message": "标准库去重完成，重复标准的规则引用已迁移。",
    }


@router.post("/standards", response_model=StandardRead)
def create_standard(payload: StandardCreate, db: Session = Depends(get_db)) -> StandardRead:
    data = payload.model_dump()
    data["clauses"] = dumps(data["clauses"])
    item = Standard(**data)
    db.add(item)
    db.commit()
    db.refresh(item)
    return StandardRead.model_validate(item, from_attributes=True).model_copy(update={"clauses": loads(item.clauses, [])})


@router.patch("/standards/{item_id}", response_model=StandardRead)
def update_standard(item_id: str, payload: StandardCreate, db: Session = Depends(get_db)) -> StandardRead:
    item = _get_or_404(db, Standard, item_id)
    data = payload.model_dump()
    data["clauses"] = dumps(data["clauses"])
    for key, value in data.items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return StandardRead.model_validate(item, from_attributes=True).model_copy(update={"clauses": loads(item.clauses, [])})


@router.delete("/standards/{item_id}", response_model=StandardRead)
def delete_standard(item_id: str, hard: bool = Query(False), db: Session = Depends(get_db)) -> StandardRead:
    item = _get_or_404(db, Standard, item_id)
    serialized = StandardRead.model_validate(item, from_attributes=True).model_copy(update={"clauses": loads(item.clauses, [])})
    if hard:
        for rule in db.scalars(select(AuditRule).where(AuditRule.standard_id == item.id)):
            rule.standard_id = None
        db.delete(item)
        db.commit()
        return serialized
    item.status = "inactive"
    db.commit()
    db.refresh(item)
    return StandardRead.model_validate(item, from_attributes=True).model_copy(update={"clauses": loads(item.clauses, [])})


@router.get("/standards/{item_id}/source")
def download_standard_source(item_id: str, db: Session = Depends(get_db)) -> FileResponse:
    item = _get_or_404(db, Standard, item_id)
    if not item.source_file:
        raise HTTPException(status_code=404, detail="Standard source file not configured")
    source_path = Path(item.source_file)
    if not source_path.exists() or not source_path.is_file():
        raise HTTPException(status_code=404, detail="Standard source file not found")
    media_type = "text/markdown" if source_path.suffix.lower() in {".md", ".txt"} else "application/octet-stream"
    return FileResponse(source_path, filename=source_path.name, media_type=media_type)


@router.get("/audit-rules", response_model=list[AuditRuleRead])
def list_audit_rules(db: Session = Depends(get_db)) -> list[AuditRule]:
    return list(db.scalars(select(AuditRule).order_by(AuditRule.created_at.desc())))


@router.post("/audit-rules", response_model=AuditRuleRead)
def create_audit_rule(payload: AuditRuleCreate, db: Session = Depends(get_db)) -> AuditRule:
    item = AuditRule(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.patch("/audit-rules/{item_id}", response_model=AuditRuleRead)
def update_audit_rule(item_id: str, payload: AuditRuleCreate, db: Session = Depends(get_db)) -> AuditRule:
    item = _get_or_404(db, AuditRule, item_id)
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return item


@router.delete("/audit-rules/{item_id}", response_model=AuditRuleRead)
def delete_audit_rule(item_id: str, hard: bool = Query(False), db: Session = Depends(get_db)) -> AuditRule:
    item = _get_or_404(db, AuditRule, item_id)
    if hard:
        db.delete(item)
        db.commit()
        return item
    item.status = "inactive"
    db.commit()
    db.refresh(item)
    return item


@router.get("/detection-items", response_model=list[DetectionItemRead])
def list_detection_items(include_inactive: bool = Query(False), db: Session = Depends(get_db)) -> list[DetectionItem]:
    statement = select(DetectionItem).order_by(DetectionItem.created_at.desc())
    if not include_inactive:
        statement = statement.where(DetectionItem.status == "active")
    return list(db.scalars(statement))


@router.post("/detection-items", response_model=DetectionItemRead)
def create_detection_item(payload: DetectionItemCreate, db: Session = Depends(get_db)) -> DetectionItem:
    item = DetectionItem(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.patch("/detection-items/{item_id}", response_model=DetectionItemRead)
def update_detection_item(item_id: str, payload: DetectionItemCreate, db: Session = Depends(get_db)) -> DetectionItem:
    item = _get_or_404(db, DetectionItem, item_id)
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return item


@router.delete("/detection-items/{item_id}", response_model=DetectionItemRead)
def delete_detection_item(item_id: str, hard: bool = Query(False), db: Session = Depends(get_db)) -> DetectionItem:
    item = _get_or_404(db, DetectionItem, item_id)
    if hard:
        db.delete(item)
        db.commit()
        return item
    item.status = "inactive"
    db.commit()
    db.refresh(item)
    return item


@router.post("/detection-items/dedupe")
def dedupe_detection_items(db: Session = Depends(get_db)) -> dict[str, Any]:
    items = list(db.scalars(select(DetectionItem).where(DetectionItem.status == "active").order_by(DetectionItem.created_at.desc())))
    seen: set[tuple[str, str, str, str, str, float, int, str]] = set()
    inactive = 0
    duplicate_groups = 0
    duplicate_keys: set[tuple[str, str, str, str, str, float, int, str]] = set()
    for item in items:
        key = (
            item.industry_id,
            item.code.strip().lower(),
            item.name.strip().lower(),
            item.method_standard.strip().lower(),
            item.judgment_standard.strip().lower(),
            float(item.price or 0),
            int(item.cycle_days or 0),
            item.package_name.strip().lower(),
        )
        if key in seen:
            item.status = "inactive"
            inactive += 1
            duplicate_keys.add(key)
        else:
            seen.add(key)
    duplicate_groups = len(duplicate_keys)
    db.commit()
    return {
        "duplicate_groups": duplicate_groups,
        "inactive": inactive,
        "message": f"已整理重复报价项目，停用 {inactive} 条重复记录。",
    }


@router.get("/labs", response_model=list[LabRead])
def list_labs(db: Session = Depends(get_db)) -> list[Lab]:
    return list(db.scalars(select(Lab).order_by(Lab.created_at.desc())))


@router.post("/labs", response_model=LabRead)
def create_lab(payload: LabCreate, db: Session = Depends(get_db)) -> Lab:
    item = Lab(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.delete("/labs/{item_id}", response_model=LabRead)
def delete_lab(item_id: str, db: Session = Depends(get_db)) -> Lab:
    item = _get_or_404(db, Lab, item_id)
    item.status = "inactive"
    db.commit()
    db.refresh(item)
    return item


@router.get("/model-providers", response_model=list[ModelProviderRead])
def list_model_providers(db: Session = Depends(get_db)) -> list[ModelProviderRead]:
    return [_serialize_model_provider(item) for item in db.scalars(select(ModelProvider).order_by(ModelProvider.created_at.desc()))]


@router.post("/model-providers", response_model=ModelProviderRead)
def create_model_provider(payload: ModelProviderCreate, db: Session = Depends(get_db)) -> ModelProviderRead:
    data = payload.model_dump()
    normalized_provider = data["provider"].strip().lower()
    normalized_model = data["model"].strip().lower()
    normalized_base_url = data.get("base_url", "").strip().rstrip("/").lower()
    item = db.scalar(
        select(ModelProvider).where(
            func.lower(ModelProvider.provider) == normalized_provider,
            func.lower(ModelProvider.model) == normalized_model,
        )
    )
    if item and item.base_url.strip().rstrip("/").lower() != normalized_base_url:
        item = None
    if item:
        existing_secret = item.api_key_secret
        existing_hint = item.api_key_hint
        for key, value in data.items():
            if key == "api_key_secret" and not str(value or "").strip():
                continue
            if key == "api_key_hint" and not str(value or "").strip() and existing_hint:
                continue
            setattr(item, key, value)
        if not item.api_key_hint and existing_secret:
            item.api_key_hint = "已保存密钥"
        item.status = "active"
    else:
        item = ModelProvider(**data)
        db.add(item)
    db.commit()
    db.refresh(item)
    return _serialize_model_provider(item)


@router.delete("/model-providers/{item_id}", response_model=ModelProviderRead)
def delete_model_provider(item_id: str, hard: bool = Query(False), db: Session = Depends(get_db)) -> ModelProviderRead:
    item = _get_or_404(db, ModelProvider, item_id)
    serialized = _serialize_model_provider(item)
    if hard:
        db.delete(item)
        db.commit()
        return serialized
    item.status = "inactive"
    db.commit()
    db.refresh(item)
    return _serialize_model_provider(item)


@router.get("/model-providers/{item_id}/secret")
def get_model_provider_secret(item_id: str, db: Session = Depends(get_db)) -> dict[str, str]:
    item = _get_or_404(db, ModelProvider, item_id)
    return {"api_key_secret": item.api_key_secret or ""}


@router.post("/model-providers/dedupe", response_model=list[ModelProviderRead])
def dedupe_model_providers(db: Session = Depends(get_db)) -> list[ModelProviderRead]:
    items = list(db.scalars(select(ModelProvider).order_by(ModelProvider.created_at.asc())))
    kept: dict[tuple[str, str, str], ModelProvider] = {}
    removed: list[ModelProvider] = []
    for item in items:
        key = (item.provider.strip().lower(), item.model.strip().lower(), item.base_url.strip().rstrip("/").lower())
        if key not in kept:
            kept[key] = item
            continue
        keeper = kept[key]
        if item.api_key_secret and not keeper.api_key_secret:
            keeper.api_key_secret = item.api_key_secret
        if item.api_key_hint and not keeper.api_key_hint:
            keeper.api_key_hint = item.api_key_hint
        keeper.supports_vision = keeper.supports_vision or item.supports_vision
        keeper.supports_json = keeper.supports_json or item.supports_json
        keeper.supports_tools = keeper.supports_tools or item.supports_tools
        keeper.default_for_text = keeper.default_for_text or item.default_for_text
        keeper.default_for_vision = keeper.default_for_vision or item.default_for_vision
        if item.status == "active":
            keeper.status = "active"
        removed.append(item)

    for item in removed:
        db.delete(item)
    db.commit()
    return [_serialize_model_provider(item) for item in db.scalars(select(ModelProvider).order_by(ModelProvider.created_at.desc()))]


@router.patch("/model-providers/{item_id}", response_model=ModelProviderRead)
def update_model_provider(item_id: str, payload: ModelProviderCreate, db: Session = Depends(get_db)) -> ModelProviderRead:
    item = _get_or_404(db, ModelProvider, item_id)
    data = payload.model_dump()
    for key, value in data.items():
        if key == "api_key_secret" and not str(value or "").strip():
            continue
        if key == "api_key_hint" and not str(value or "").strip() and item.api_key_hint:
            continue
        setattr(item, key, value)
    if not item.api_key_hint and item.api_key_secret:
        item.api_key_hint = "已保存密钥"
    db.commit()
    db.refresh(item)
    return _serialize_model_provider(item)


def _serialize_model_provider(item: ModelProvider) -> ModelProviderRead:
    api_key_saved = bool((item.api_key_secret or "").strip())
    return ModelProviderRead.model_validate(item, from_attributes=True).model_copy(
        update={
            "api_key_saved": api_key_saved,
            "api_key_hint": item.api_key_hint or ("已保存密钥" if api_key_saved else ""),
        }
    )


def _serialize_import_task(item: ImportTask) -> ImportTaskRead:
    return ImportTaskRead.model_validate(item, from_attributes=True).model_copy(
        update={"parsed_result": loads(item.parsed_result, {})}
    )


@router.get("/import-tasks", response_model=list[ImportTaskRead])
def list_import_tasks(
    target_library: Optional[str] = Query(None),
    db: Session = Depends(get_db),
) -> list[ImportTaskRead]:
    statement = select(ImportTask).order_by(ImportTask.created_at.desc())
    if target_library:
        statement = statement.where(ImportTask.target_library == target_library)
    return [_serialize_import_task(item) for item in db.scalars(statement)]


@router.get("/import-tasks/{task_id}", response_model=ImportTaskRead)
def get_import_task(task_id: str, db: Session = Depends(get_db)) -> ImportTaskRead:
    item = _get_or_404(db, ImportTask, task_id)
    return _serialize_import_task(item)


@router.delete("/import-tasks/{task_id}", response_model=ImportTaskRead)
def delete_import_task(task_id: str, db: Session = Depends(get_db)) -> ImportTaskRead:
    item = _get_or_404(db, ImportTask, task_id)
    serialized = _serialize_import_task(item)
    if item.file_path:
        saved_file = Path(item.file_path)
        if saved_file.exists() and saved_file.is_file():
            saved_file.unlink()
    db.delete(item)
    db.commit()
    return serialized


@router.post("/standard-rule-library/import")
@router.post("/imports/standards")
async def import_standard_rule_library(
    industry_id: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    industry = db.get(Industry, industry_id)
    if not industry:
        raise HTTPException(status_code=404, detail="Industry not found")
    content = await file.read()
    saved_path = _save_import_file(file.filename or "standard-rule-library", content)
    rows = _read_tabular_file(file.filename or "", content)
    if not rows:
        parsed_document = get_document_parser().parse(str(saved_path), content)
        text = str(parsed_document.get("text") or _safe_text(content))
        provider = _default_text_model(db)
        parsed_library = get_model_gateway().parse_library_document(
            provider=provider,
            target_library="standard_rule",
            industry_name=industry.name,
            filename=file.filename or "",
            text=text,
        )
        standards_created, rules_created, clause_chunks_created = _create_standards_and_rules_from_parsed(
            db,
            industry_id=industry_id,
            source_file=str(saved_path),
            fallback_filename=file.filename or "导入标准文档",
            parsed_library=parsed_library,
            source_text=text,
        )
        import_task = _record_import_task(
            db,
            filename=file.filename or "",
            file_path=str(saved_path),
            target_library="standard_rule",
            status="needs_review",
            model=str(parsed_library.get("model") or "local-library-parser"),
            parsed_result={
                "standards_created": standards_created,
                "rules_created": rules_created,
                "clause_chunks_created": clause_chunks_created,
                "requires_manual_confirmation": True,
                "document_parser": parsed_document.get("provider"),
                "model": parsed_library.get("model"),
                "model_note": parsed_library.get("model_note", ""),
                "source_excerpt": text[:1200],
                "message": "已从原文文档解析标准/规则并标记为需人工确认。",
            },
        )
        return {
            "standards_created": standards_created,
            "rules_created": rules_created,
            "clause_chunks_created": clause_chunks_created,
            "import_task": _serialize_import_task(import_task).model_dump(mode="json"),
            "message": f"已导入标准规则库文档，并生成 {clause_chunks_created} 个全文条款切片。",
        }

    standards_created = 0
    rules_created = 0
    standard_by_code: dict[str, Standard] = {}
    for row in rows:
        code = _pick(row, ["标准编号", "标准号", "code", "standard_code"]) or _filename_code(file.filename or "STANDARD")
        name = _pick(row, ["标准名称", "名称", "name", "standard_name"]) or Path(file.filename or code).stem
        version = _pick(row, ["版本", "version"]) or "现行"
        effective_date = _pick(row, ["生效日期", "effective_date"]) or ""
        trigger = _pick(row, ["触发条件", "检查要求", "条款", "trigger", "clause"]) or ""
        field_key = _pick(row, ["字段键", "字段", "field_key", "field"]) or "claims"
        risk_level = _normalize_risk(_pick(row, ["风险等级", "risk_level", "risk"]) or "medium")
        suggestion = _pick(row, ["整改建议", "建议", "suggestion"]) or "建议结合导入标准进行人工复核。"

        standard = standard_by_code.get(code)
        if not standard:
            standard = _find_standard_by_code(db, industry_id, code)
            if standard:
                standard.name = name or standard.name
                standard.version = version or standard.version
                standard.effective_date = effective_date or standard.effective_date
                standard.source_file = str(saved_path)
                standard.status = "active"
            else:
                standard = Standard(
                    industry_id=industry_id,
                    code=code,
                    name=name,
                    version=version,
                    effective_date=effective_date,
                    source_file=str(saved_path),
                    clauses=dumps([]),
                )
                db.add(standard)
                db.flush()
                standards_created += 1
            standard_by_code[code] = standard

        if trigger:
            db.add(
                AuditRule(
                    industry_id=industry_id,
                    standard_id=standard.id,
                    name=_pick(row, ["规则名称", "规则", "rule_name"]) or f"{code} - {field_key}",
                    rule_type=_pick(row, ["规则类型", "rule_type"]) or "ai",
                    field_key=field_key,
                    trigger=trigger,
                    risk_level=risk_level,
                    suggestion=suggestion,
                )
            )
            rules_created += 1
    db.commit()
    import_task = _record_import_task(
        db,
        filename=file.filename or "",
        file_path=str(saved_path),
        target_library="standard_rule",
        status="completed",
        parsed_result={
            "standards_created": standards_created,
            "rules_created": rules_created,
            "requires_manual_confirmation": True,
            "message": "已从表格解析出标准依据和审核规则。",
        },
    )
    return {
        "standards_created": standards_created,
        "rules_created": rules_created,
        "import_task": _serialize_import_task(import_task).model_dump(mode="json"),
        "message": "标准规则库已导入。",
    }


@router.post("/standard-rule-library/import-url")
async def import_standard_rule_library_url(
    payload: dict[str, str],
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    industry_id = payload.get("industry_id", "")
    url = payload.get("url", "").strip()
    if not url.startswith(("http://", "https://")):
        raise HTTPException(status_code=400, detail="URL must start with http:// or https://")
    industry = db.get(Industry, industry_id)
    if not industry:
        raise HTTPException(status_code=404, detail="Industry not found")
    try:
        with httpx.Client(timeout=45, follow_redirects=True, headers={"User-Agent": "Mozilla/5.0"}) as client:
            response = client.get(url)
            response.raise_for_status()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"URL download failed: {exc}") from exc

    filename = _filename_from_url(url, response.headers.get("content-type", ""))
    saved_path = _save_import_file(filename, response.content)
    parsed_document = get_document_parser().parse(str(saved_path), response.content)
    text = str(parsed_document.get("text") or _safe_text(response.content))
    parsed_library = _parse_standard_identity_from_text(text, fallback_filename=filename)
    standards_created, rules_created, clause_chunks_created = _create_standards_and_rules_from_parsed(
        db,
        industry_id=industry_id,
        source_file=str(saved_path),
        fallback_filename=filename,
        parsed_library=parsed_library,
        source_text=text,
    )
    import_task = _record_import_task(
        db,
        filename=filename,
        file_path=str(saved_path),
        target_library="standard_rule",
        status="needs_review",
        model="url-fulltext-importer",
        parsed_result={
            "source_url": url,
            "standards_created": standards_created,
            "rules_created": rules_created,
            "clause_chunks_created": clause_chunks_created,
            "requires_manual_confirmation": True,
            "document_parser": parsed_document.get("provider"),
            "source_excerpt": text[:1200],
            "message": "已从 URL 下载原文并生成全文条款切片。",
        },
    )
    return {
        "standards_created": standards_created,
        "rules_created": rules_created,
        "clause_chunks_created": clause_chunks_created,
        "import_task": _serialize_import_task(import_task).model_dump(mode="json"),
        "message": f"已从 URL 导入原文，并生成 {clause_chunks_created} 个全文条款切片。",
    }


@router.post("/quote-library/import")
async def import_quote_library(
    industry_id: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    industry = db.get(Industry, industry_id)
    if not industry:
        raise HTTPException(status_code=404, detail="Industry not found")
    content = await file.read()
    saved_path = _save_import_file(file.filename or "quote-library", content)
    rows = _read_tabular_file(file.filename or "", content)
    if not rows:
        parsed_document = get_document_parser().parse(str(saved_path), content)
        text = str(parsed_document.get("text") or _safe_text(content))
        provider = _default_text_model(db)
        parsed_library = get_model_gateway().parse_library_document(
            provider=provider,
            target_library="quote",
            industry_name=industry.name,
            filename=file.filename or "",
            text=text,
        )
        created = _create_quote_items_from_parsed(
            db,
            industry_id=industry_id,
            parsed_library=parsed_library,
        )
        import_task = _record_import_task(
            db,
            filename=file.filename or "",
            file_path=str(saved_path),
            target_library="quote",
            status="needs_review",
            model=str(parsed_library.get("model") or "local-library-parser"),
            parsed_result={
                "items_created": created,
                "requires_manual_confirmation": True,
                "document_parser": parsed_document.get("provider"),
                "model": parsed_library.get("model"),
                "model_note": parsed_library.get("model_note", ""),
                "source_excerpt": text[:1200],
                "message": "已保存报价原文并尝试解析报价项目，解析结果需人工确认。",
            },
        )
        return {
            "items_created": created,
            "import_task": _serialize_import_task(import_task).model_dump(mode="json"),
            "message": "已导入项目报价库文档，并生成需人工确认的解析结果。",
        }

    created = 0
    for row in rows:
        name = _pick(row, ["项目名称", "检测项目", "name", "item"]) or ""
        if not name:
            continue
        code = _pick(row, ["编号", "项目编号", "code"]) or f"IMP{created + 1:03d}"
        price = _to_float(_pick(row, ["价格", "报价", "price", "amount"]) or "0")
        cycle_days = int(_to_float(_pick(row, ["周期", "周期/天", "cycle_days", "days"]) or "5") or 5)
        db.add(
            DetectionItem(
                industry_id=industry_id,
                code=code,
                name=name,
                method_standard=_pick(row, ["方法标准", "检测标准", "method_standard", "standard"]) or "",
                judgment_standard=_pick(row, ["判定标准", "judgment_standard"]) or "",
                price=price,
                cycle_days=cycle_days,
                sample_amount=_pick(row, ["样品量", "样品/原料要求", "原料要求", "sample_amount"]) or "",
                package_name=_pick(row, ["套餐", "package_name"]) or "导入报价库",
            )
        )
        created += 1
    db.commit()
    import_task = _record_import_task(
        db,
        filename=file.filename or "",
        file_path=str(saved_path),
        target_library="quote",
        status="completed",
        parsed_result={
            "items_created": created,
            "requires_manual_confirmation": True,
            "message": "已从表格解析出项目报价条目。",
        },
    )
    return {
        "items_created": created,
        "import_task": _serialize_import_task(import_task).model_dump(mode="json"),
        "message": "项目报价库已导入。",
    }


def _record_import_task(
    db: Session,
    filename: str,
    file_path: str,
    target_library: str,
    status: str,
    parsed_result: dict[str, Any],
    error_message: str = "",
    model: str = "local-parser-v1",
) -> ImportTask:
    item = ImportTask(
        filename=filename,
        file_path=file_path,
        file_type=Path(filename).suffix.lower().lstrip("."),
        target_library=target_library,
        status=status,
        model=model,
        parsed_result=dumps(parsed_result),
        error_message=error_message,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


def _default_text_model(db: Session) -> Optional[ModelProvider]:
    return db.scalar(
        select(ModelProvider).where(ModelProvider.default_for_text.is_(True), ModelProvider.status == "active")
    ) or db.scalar(select(ModelProvider).where(ModelProvider.status == "active").order_by(ModelProvider.created_at.desc()))


def _create_standards_and_rules_from_parsed(
    db: Session,
    industry_id: str,
    source_file: str,
    fallback_filename: str,
    parsed_library: dict[str, Any],
    source_text: str = "",
) -> tuple[int, int, int]:
    raw_standards = parsed_library.get("standards")
    standards_payload = raw_standards if isinstance(raw_standards, list) else []
    raw_rules = parsed_library.get("rules")
    rules_payload = raw_rules if isinstance(raw_rules, list) else []
    if not standards_payload:
        standards_payload = [
            {
                "code": _filename_code(fallback_filename),
                "name": Path(fallback_filename).stem or "导入标准文档",
                "version": "导入文档",
                "effective_date": "",
                "clauses": [],
            }
        ]

    created_standards = 0
    created_rules = 0
    created_clause_chunks = 0
    standard_by_code: dict[str, Standard] = {}
    for raw in standards_payload:
        if not isinstance(raw, dict):
            continue
        code = str(raw.get("code") or _filename_code(fallback_filename))[:80]
        clauses = raw.get("clauses") if isinstance(raw.get("clauses"), list) else []
        standard = _find_standard_by_code(db, industry_id, code)
        if standard:
            standard.name = str(raw.get("name") or standard.name or code)[:160]
            standard.version = str(raw.get("version") or standard.version or "导入文档")[:40]
            standard.effective_date = str(raw.get("effective_date") or standard.effective_date or "")[:20]
            standard.source_file = source_file
            standard.clauses = dumps(clauses or loads(standard.clauses, []))
            standard.status = "active"
        else:
            standard = Standard(
                industry_id=industry_id,
                code=code,
                name=str(raw.get("name") or Path(fallback_filename).stem or code)[:160],
                version=str(raw.get("version") or "导入文档")[:40],
                effective_date=str(raw.get("effective_date") or "")[:20],
                source_file=source_file,
                clauses=dumps(clauses),
                status="active",
            )
            db.add(standard)
            db.flush()
            created_standards += 1
        standard_by_code[code] = standard

    if source_text.strip():
        targets = list(standard_by_code.values())
        if not targets:
            fallback = Standard(
                industry_id=industry_id,
                code=_filename_code(fallback_filename),
                name=Path(fallback_filename).stem or "导入标准文档",
                version="导入文档",
                source_file=source_file,
                clauses=dumps([]),
                status="active",
            )
            db.add(fallback)
            db.flush()
            targets = [fallback]
            created_standards += 1
        for standard in targets[:1]:
            created_clause_chunks += _replace_standard_clauses(db, standard, source_text, source_file)

    default_standard = next(iter(standard_by_code.values()), None)
    for raw in rules_payload:
        if not isinstance(raw, dict):
            continue
        standard = standard_by_code.get(str(raw.get("standard_code") or "")) or default_standard
        trigger = str(raw.get("trigger") or raw.get("source_excerpt") or "")
        if not trigger:
            continue
        db.add(
            AuditRule(
                industry_id=industry_id,
                standard_id=standard.id if standard else None,
                name=str(raw.get("name") or f"{standard.code if standard else '导入'} - 规则")[:160],
                rule_type=str(raw.get("rule_type") or "ai")[:40],
                field_key=str(raw.get("field_key") or "claims")[:80],
                trigger=trigger,
                risk_level=_normalize_risk(str(raw.get("risk_level") or "medium")),
                suggestion=str(raw.get("suggestion") or "请结合来源文件和现行标准人工复核。"),
                status="active",
            )
        )
        created_rules += 1

    db.commit()
    return created_standards, created_rules, created_clause_chunks


def _create_quote_items_from_parsed(
    db: Session,
    industry_id: str,
    parsed_library: dict[str, Any],
) -> int:
    raw_items = parsed_library.get("quote_items") or parsed_library.get("items")
    items_payload = raw_items if isinstance(raw_items, list) else []
    created = 0
    for raw in items_payload:
        if not isinstance(raw, dict):
            continue
        name = str(raw.get("name") or raw.get("item") or "").strip()
        if not name:
            continue
        db.add(
            DetectionItem(
                industry_id=industry_id,
                code=str(raw.get("code") or f"IMP{created + 1:03d}")[:60],
                name=name[:160],
                method_standard=str(raw.get("method_standard") or raw.get("standard") or "")[:120],
                judgment_standard=str(raw.get("judgment_standard") or "")[:120],
                price=_to_float(str(raw.get("price") or raw.get("amount") or "0")),
                cycle_days=int(_to_float(str(raw.get("cycle_days") or raw.get("days") or "5")) or 5),
                sample_amount=str(raw.get("sample_amount") or raw.get("sample_requirement") or raw.get("remark") or "")[:80],
                package_name=str(raw.get("package_name") or "导入报价库")[:100],
                status="active",
            )
        )
        created += 1
    db.commit()
    return created


def _save_import_file(filename: str, content: bytes) -> Path:
    settings = get_settings()
    suffix = Path(filename).suffix
    safe_stem = Path(filename).stem.replace("/", "_").replace("\\", "_")[:80] or "import"
    path = settings.upload_dir / f"import-{uuid4().hex}-{safe_stem}{suffix}"
    path.write_bytes(content)
    return path


def _filename_from_url(url: str, content_type: str) -> str:
    parsed = urlparse(url)
    name = unquote(Path(parsed.path).name or "downloaded-standard")
    if "." not in name:
        if "pdf" in content_type:
            name += ".pdf"
        elif "word" in content_type or "officedocument" in content_type:
            name += ".docx"
        elif "html" in content_type:
            name += ".html"
        else:
            name += ".txt"
    return name[:160]


def _parse_standard_identity_from_text(text: str, fallback_filename: str) -> dict[str, Any]:
    head = "\n".join((text or "").splitlines()[:40])
    code_match = re.search(r"\b(GB(?:/T)?\s*\d+(?:\.\d+)?-\d{4}|CNCA[-—－][A-Z0-9-]+:\d{4})\b", head, re.I)
    code = code_match.group(1).replace(" ", "") if code_match else _filename_code(fallback_filename)
    title = ""
    for line in head.splitlines():
        clean = line.strip()
        if len(clean) >= 6 and any(keyword in clean for keyword in ["标准", "办法", "规定", "规则", "规范"]):
            title = clean[:160]
            break
    if not title:
        title = Path(fallback_filename).stem
    version_match = re.search(r"(\d{4})", code)
    return {
        "standards": [
            {
                "code": code,
                "name": title,
                "version": version_match.group(1) if version_match else "导入文档",
                "effective_date": "",
                "clauses": [],
            }
        ],
        "rules": [],
    }


def _read_tabular_file(filename: str, content: bytes) -> list[dict[str, str]]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        return _rows_to_dicts(rows)
    text = _safe_text(content)
    if suffix == ".csv" or "," in text.splitlines()[0:1][0] if text.splitlines() else False:
        reader = csv.DictReader(io.StringIO(text))
        return [{str(k).strip(): str(v or "").strip() for k, v in row.items() if k} for row in reader]
    lines = [line for line in text.splitlines() if line.strip()]
    delimiter = "\t" if any("\t" in line for line in lines[:3]) else None
    if delimiter:
        return _rows_to_dicts([line.split(delimiter) for line in lines])
    return []


def _rows_to_dicts(rows: list[Any]) -> list[dict[str, str]]:
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    result = []
    for row in rows[1:]:
        item = {headers[index]: str(value or "").strip() for index, value in enumerate(row) if index < len(headers) and headers[index]}
        if any(item.values()):
            result.append(item)
    return result


def _pick(row: dict[str, str], keys: list[str]) -> str:
    normalized = {key.strip().lower(): value for key, value in row.items()}
    for key in keys:
        value = normalized.get(key.strip().lower())
        if value:
            return value.strip()
    return ""


def _safe_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return ""


def _filename_code(filename: str) -> str:
    stem = Path(filename).stem.upper().replace(" ", "-")
    return stem[:60] or "IMPORTED-STANDARD"


def _normalize_risk(value: str) -> str:
    lowered = value.lower()
    if "高" in value or lowered == "high":
        return "high"
    if "低" in value or lowered == "low":
        return "low"
    return "medium"


def _to_float(value: str) -> float:
    try:
        return float("".join(ch for ch in value if ch.isdigit() or ch == ".") or 0)
    except ValueError:
        return 0
